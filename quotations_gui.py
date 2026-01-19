import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import datetime as _dt
import os
from typing import Callable, List, Dict, Any


def run_quotations(parent: tk.Widget, list_boards: Callable[[], List[Dict[str, Any]]]):
    """
    Build a Quotations page inside the given parent widget.

    - Left: Boards from database with basic filters and selection
    - Right: Current quotation items
    - Actions: Add to quotation, remove, clear, and Export to Excel (.xlsx) with CSV fallback
    """
    # Root container for this page
    page = ttk.Frame(parent)
    page.pack(fill="both", expand=True)

    style = ttk.Style(page)
    try:
        style.theme_use("clam")
    except Exception:
        pass

    # Layout: two columns
    page.columnconfigure(0, weight=1)
    page.columnconfigure(1, weight=1)
    page.rowconfigure(1, weight=1)

    ttk.Label(page, text="Quotations", font=("Segoe UI", 16, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", padx=10, pady=(10, 6))

    # Left: boards list
    left = ttk.LabelFrame(page, text="Boards")
    left.grid(row=1, column=0, sticky="nsew", padx=(10, 5), pady=(0, 10))

    # Search + Filter + Sort bar
    search_var = tk.StringVar(value="")
    sort_var = tk.StringVar(value="Running No Right (Asc)")
    site_var = tk.StringVar(value="All")
    size_var = tk.StringVar(value="All")
    # Month filter state for Date Request
    month_names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    month_vars = {m: tk.BooleanVar(value=False) for m in month_names}
    bar = ttk.Frame(left)
    bar.pack(fill="x", padx=8, pady=6)
    ttk.Label(bar, text="Search:").pack(side="left")
    ent_search = ttk.Entry(bar, textvariable=search_var, width=28)
    ent_search.pack(side="left", padx=6)
    ttk.Button(bar, text="Filter...", command=lambda: open_filters_dialog()).pack(side="left", padx=6)
    ttk.Button(bar, text="Sort...", command=lambda: open_sort_dialog()).pack(side="left", padx=6)

    # Tree for boards
    # Multi-select via checkbox column
    selected_ids: set[str] = set()
    cols = ("Select", "Board ID", "Site Name", "Running No Right", "Size")
    tv_boards = ttk.Treeview(left, columns=cols, show="headings", height=12, selectmode="extended")
    def toggle_select_all():
        iids = tv_boards.get_children()
        all_ids = {iid.split(":", 1)[1] if ":" in iid else iid for iid in iids}
        if len(selected_ids) == len(all_ids) and len(all_ids) > 0:
            selected_ids.clear()
        else:
            selected_ids.clear(); selected_ids.update(all_ids)
        refresh_boards()
    for c in cols:
        if c == "Select":
            tv_boards.heading(c, text=c, command=toggle_select_all)
            w = 60
        else:
            tv_boards.heading(c, text=c)
            w = 100
            if c == "Site Name":
                w = 180
            if c == "Running No Right":
                w = 120
        tv_boards.column(c, width=w, stretch=False)
    tv_boards.pack(fill="both", expand=True, padx=8, pady=(0, 8))

    # Right: quotation items
    right = ttk.LabelFrame(page, text="Quotation Items")
    right.grid(row=1, column=1, sticky="nsew", padx=(5, 10), pady=(0, 10))

    # Quotation list includes Issue and Quantity (editable)
    q_cols = ("Board ID", "Module No", "RN No", "Issue", "Quantity")
    tv_quote = ttk.Treeview(right, columns=q_cols, show="headings", height=12, selectmode="extended")
    for c in q_cols:
        tv_quote.heading(c, text=c)
        w = 100
        if c in ("Issue",):
            w = 140
        if c in ("Module No", "RN No"):
            w = 90
        tv_quote.column(c, width=w, stretch=False)
    tv_quote.pack(fill="both", expand=True, padx=8, pady=(0, 8))

    # Button row
    actions = ttk.Frame(page)
    actions.grid(row=2, column=0, columnspan=2, sticky="ew", padx=10, pady=(0, 10))
    actions.columnconfigure(0, weight=1)
    actions.columnconfigure(1, weight=1)

    btn_add = ttk.Button(actions, text="Add Selected →")
    btn_remove = ttk.Button(actions, text="← Remove Selected")
    btn_clear = ttk.Button(actions, text="Clear")
    btn_export = ttk.Button(actions, text="Export…")
    btn_add.grid(row=0, column=0, sticky="w", padx=4)
    btn_remove.grid(row=0, column=0, sticky="e", padx=4)
    btn_clear.grid(row=0, column=1, sticky="w", padx=4)
    btn_export.grid(row=0, column=1, sticky="e", padx=4)

    # Helpers
    def _matches(b: Dict[str, Any], q: str) -> bool:
        if not q:
            return True
        ql = q.lower()
        for k in ("board_id", "name", "running_no", "size", "running_no_p1", "running_no_p2"):
            v = b.get(k)
            if v and ql in str(v).lower():
                return True
        return False

    def unique_values(key: str):
        try:
            vals = sorted({(str(b.get(key)) or "") for b in list_boards() if b.get(key)})
            return ["All"] + vals
        except Exception:
            return ["All"]

    def open_filters_dialog():
        win = tk.Toplevel(page)
        win.title("Filters")
        frm = ttk.Frame(win); frm.pack(fill="both", expand=True, padx=10, pady=10)
        ttk.Label(frm, text="Site:").grid(row=0, column=0, sticky="w", padx=6, pady=4)
        ttk.Label(frm, text="Size:").grid(row=0, column=2, sticky="w", padx=6, pady=4)
        site_cb = ttk.Combobox(frm, width=24, textvariable=site_var, state="readonly", values=unique_values("name"))
        size_cb = ttk.Combobox(frm, width=12, textvariable=size_var, state="readonly", values=unique_values("size"))
        site_cb.grid(row=0, column=1, sticky="w", padx=6, pady=4)
        size_cb.grid(row=0, column=3, sticky="w", padx=6, pady=4)
        # Month checkboxes by Date Request
        ttk.Label(frm, text="Months (Date Request):").grid(row=1, column=0, sticky="w", padx=6, pady=4)
        months_frame = ttk.Frame(frm)
        months_frame.grid(row=1, column=1, columnspan=3, sticky="w", padx=6, pady=4)
        for idx, m in enumerate(month_names):
            r = 0 if idx < 6 else 1
            c = idx if idx < 6 else idx-6
            ttk.Checkbutton(months_frame, text=m, variable=month_vars[m]).grid(row=r, column=c, padx=4, pady=2, sticky="w")
        def select_all_months():
            for v in month_vars.values(): v.set(True)
        def clear_all_months():
            for v in month_vars.values(): v.set(False)
        ttk.Button(months_frame, text="All", command=select_all_months).grid(row=2, column=0, padx=4, pady=2, sticky="w")
        ttk.Button(months_frame, text="None", command=clear_all_months).grid(row=2, column=1, padx=4, pady=2, sticky="w")
        def on_ok():
            refresh_boards(); win.destroy()
        ttk.Button(frm, text="OK", command=on_ok).grid(row=1, column=3, sticky="e", padx=6, pady=8)
        ttk.Button(frm, text="Clear", command=lambda: [site_var.set("All"), size_var.set("All"), clear_all_months(), refresh_boards()]).grid(row=1, column=2, sticky="e", padx=6, pady=8)

    def open_sort_dialog():
        win = tk.Toplevel(page)
        win.title("Sort")
        frm = ttk.Frame(win); frm.pack(fill="both", expand=True, padx=10, pady=10)
        ttk.Label(frm, text="Sort by:").grid(row=0, column=0, sticky="w", padx=6, pady=4)
        options = [
            "Running No Right (Asc)",
            "Running No Right (Desc)",
            "Date Request (Newest)",
            "Date Request (Oldest)",
            "Module Number (Asc)",
            "Module Number (Desc)",
            "Site Name (A-Z)",
        ]
        cb = ttk.Combobox(frm, width=26, textvariable=sort_var, state="readonly", values=options)
        cb.grid(row=0, column=1, sticky="w", padx=6, pady=4)
        ttk.Button(frm, text="OK", command=lambda: [refresh_boards(), win.destroy()]).grid(row=1, column=1, sticky="e", padx=6, pady=8)

    def refresh_boards():
        tv_boards.delete(*tv_boards.get_children())
        query = search_var.get().strip()
        try:
            data = list_boards()
        except Exception as e:
            messagebox.showerror("Error", f"Unable to list boards: {e}")
            data = []
        # Apply filters
        s = site_var.get()
        if s and s != "All":
            data = [b for b in data if str(b.get("name")) == s]
        sz = size_var.get()
        if sz and sz != "All":
            data = [b for b in data if str(b.get("size")) == sz]
        # Months by date_request
        if month_vars:
            sel_months = [m for m, var in month_vars.items() if var.get()]
            if sel_months:
                mindex = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,"Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
                nums = {mindex[m] for m in sel_months}
                def dm(b):
                    d = b.get("date_request")
                    if not d:
                        return None
                    try:
                        return int(str(d).split("-")[1])
                    except Exception:
                        return None
                data = [b for b in data if dm(b) in nums]
        # Apply search
        data = [b for b in data if _matches(b, query)]
        # Sort
        def keydate(b):
            import datetime as _dt
            try:
                return _dt.datetime.strptime(str(b.get("date_request")), "%Y-%m-%d")
            except Exception:
                return _dt.datetime(1900, 1, 1)
        def keyint(b, k):
            try:
                return int(str(b.get(k) or "0"))
            except Exception:
                return 0
        if sort_var.get() == "Running No Right (Asc)":
            data.sort(key=lambda b: keyint(b, "running_no_p2"))
        elif sort_var.get() == "Running No Right (Desc)":
            data.sort(key=lambda b: keyint(b, "running_no_p2"), reverse=True)
        elif sort_var.get() == "Date Request (Newest)":
            data.sort(key=keydate, reverse=True)
        elif sort_var.get() == "Date Request (Oldest)":
            data.sort(key=keydate)
        elif sort_var.get() == "Module Number (Asc)":
            data.sort(key=lambda b: keyint(b, "module_number"))
        elif sort_var.get() == "Module Number (Desc)":
            data.sort(key=lambda b: keyint(b, "module_number"), reverse=True)
        elif sort_var.get() == "Site Name (A-Z)":
            data.sort(key=lambda b: str(b.get("name") or ""))
        # Populate
        for b in data:
            bid = str(b.get("board_id"))
            chk = "☑" if bid in selected_ids else "☐"
            rn_right = str(b.get("running_no_p2") or "") or (str(b.get("running_no") or ""))
            rn_right = rn_right if rn_right else "-"
            tv_boards.insert("", "end", iid=f"b:{bid}", values=(chk, bid, b.get("name") or "", rn_right, b.get("size") or ""))

    def _get_board_by_id(board_id: str) -> Dict[str, Any] | None:
        try:
            for b in list_boards():
                if str(b.get("board_id")) == str(board_id):
                    return b
        except Exception:
            return None
        return None

    def add_selected_to_quote():
        # Prefer checkbox selections; fallback to row selection
        rows_to_add = []
        if selected_ids:
            for bid in list(selected_ids):
                iid = f"b:{bid}"
                if tv_boards.exists(iid):
                    rows_to_add.append(tv_boards.item(iid, "values"))
        else:
            sel = tv_boards.selection()
            for iid in sel:
                rows_to_add.append(tv_boards.item(iid, "values"))
        if not rows_to_add:
            messagebox.showinfo("Add", "Select one or more boards to add.")
            return
        for vals in rows_to_add:
            # vals: (chk, bid, site, rn_right, size)
            bid = str(vals[1])
            b = _get_board_by_id(bid) or {}
            module_no = b.get("module_number") or "-"
            rn_right = vals[3]
            issue = ""  # admin can fill later
            qty = 1
            tv_quote.insert("", "end", values=(bid, module_no, rn_right, issue, qty))

    def remove_selected_from_quote():
        sel = tv_quote.selection()
        for iid in sel:
            tv_quote.delete(iid)

    def clear_quote():
        tv_quote.delete(*tv_quote.get_children())

    def export_quote():
        # Gather rows
        rows = [tv_quote.item(i, "values") for i in tv_quote.get_children()]
        if not rows:
            messagebox.showinfo("Export", "Quotation list is empty.")
            return
        # Prompt for metadata: quotation id, project name/code, modules code, total repair modules, date request
        def prompt_meta(defaults: dict) -> dict | None:
            win = tk.Toplevel(page); win.title("Quotation Details")
            frm = ttk.Frame(win); frm.pack(fill="both", expand=True, padx=10, pady=10)
            fields = [
                ("Quotation ID", "quotation_id"),
                ("Project Name", "project_name"),
                ("Project Code", "project_code"),
                ("Modules Code", "modules_code"),
                ("Total Repair Modules", "total_repair_modules"),
                ("Date Request (dd/mm/yyyy)", "date_request"),
                ("Pixel", "pixel"),
            ]
            vars: dict[str, tk.StringVar] = {}
            for r, (label, key) in enumerate(fields):
                ttk.Label(frm, text=label+":").grid(row=r, column=0, sticky="w", padx=6, pady=4)
                v = tk.StringVar(value=str(defaults.get(key, "")))
                ttk.Entry(frm, textvariable=v, width=30).grid(row=r, column=1, sticky="w", padx=6, pady=4)
                vars[key] = v
            result: dict | None = {k: v.get().strip() for k, v in vars.items()}
            def ok():
                nonlocal result
                result = {k: v.get().strip() for k, v in vars.items()}
                win.destroy()
            def cancel():
                nonlocal result
                result = None
                win.destroy()
            btns = ttk.Frame(frm); btns.grid(row=len(fields), column=0, columnspan=2, sticky="e")
            ttk.Button(btns, text="Cancel", command=cancel).pack(side="left", padx=6)
            ttk.Button(btns, text="OK", command=ok).pack(side="left", padx=6)
            win.grab_set(); page.wait_window(win)
            return result

        computed_total = 0
        for r in rows:
            try:
                computed_total += int(r[4])
            except Exception:
                pass
        # Derive a sensible default for Pixel from the first selected board
        try:
            first_bid = rows[0][0]
            b0 = _get_board_by_id(str(first_bid)) or {}
            default_pixel = str(b0.get('pixel') or b0.get('size') or '')
        except Exception:
            default_pixel = ''

        defaults = {
            "quotation_id": "1",
            "project_name": "",
            "project_code": "",
            "modules_code": "",
            "total_repair_modules": str(computed_total or len(rows)),
            "date_request": _dt.date.today().strftime('%d/%m/%Y'),
            "pixel": default_pixel,
        }
        meta = prompt_meta(defaults)
        if meta is None:
            return
        # Choose file
        path = filedialog.asksaveasfilename(
            title="Export Quotation",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", ".xlsx"), ("CSV", ".csv")],
        )
        if not path:
            return
        if path.lower().endswith(".xlsx"):
            try:
                export_to_xlsx(path, rows, meta)
                messagebox.showinfo("Export", f"Saved to {path}")
                return
            except Exception as e:
                messagebox.showwarning("Excel export failed", f"Falling back to CSV. Error: {e}")
                # Fall through to CSV
        try:
            export_to_csv(path if path.lower().endswith('.csv') else path + '.csv', rows, meta)
            messagebox.showinfo("Export", f"Saved to {path if path.lower().endswith('.csv') else path + '.csv'}")
        except Exception as e:
            messagebox.showerror("Export", f"Failed to export: {e}")

    def export_to_xlsx(path: str, rows, meta: dict):
        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # type: ignore
        except Exception as e:
            raise RuntimeError("openpyxl is required for .xlsx export") from e
        wb = Workbook()
        # Common alignment for table cells
        align_center = Alignment(horizontal="center", vertical="center")
        # Split into multiple pages of 10 boards each
        page_size = 10
        pages = [rows[i:i+page_size] for i in range(0, len(rows), page_size)] or [[]]
        from openpyxl.utils import get_column_letter  # type: ignore
        # Try to load a logo image if available
        logo_path_candidates = [
            os.path.join(os.path.dirname(__file__), 'assets', 'IDS LOGO.png'),
            os.path.join(os.path.dirname(__file__), 'assets', 'logo.png'),
            os.path.join(os.path.dirname(__file__), 'assets', 'logo.jpg'),
        ]
        logo_path = next((p for p in logo_path_candidates if os.path.exists(p)), None)
        # Create a single sheet with all tables stacked vertically
        ws = wb.active
        ws.title = "Quotation"
        # Hide default Excel gridlines to match printed look
        ws.sheet_view.showGridLines = False
        # Fit to A4 portrait and narrow margins to squeeze content
        try:
            ws.page_setup.orientation = 'portrait'
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
        except Exception:
            pass
        try:
            ws.page_margins.left = 0.3
            ws.page_margins.right = 0.3
            ws.page_margins.top = 0.5
            ws.page_margins.bottom = 0.5
        except Exception:
            pass

        # Helper to apply borders to a merged range (outer box)
        def apply_range_border(r1, c1, r2, c2, border):
            for rr in range(r1, r2 + 1):
                for cc in range(c1, c2 + 1):
                    ws.cell(row=rr, column=cc).border = border

        thin = Side(style="thin", color="000000")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        for page_index, page_rows in enumerate(pages, start=1):
            # Start position for this table
            if page_index == 1:
                r = 1
                table_start_row = r
            else:
                table_start_row = ws.max_row + 1
                r = table_start_row
            
            # Add logo and company info to every table
            logo_img = None
            if logo_path:
                try:
                    from openpyxl.drawing.image import Image as XLImage  # type: ignore
                    img = XLImage(logo_path)
                    img.height = 80  # reduced height
                    img.width = 400  # temporary; adjusted after column widths are set
                    ws.add_image(img, f"A{r}")
                    logo_img = img
                except Exception:
                    logo_img = None
            # Company name line
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
            c = ws.cell(row=r, column=2, value="IDS BEYOND MEDIA SDN BHD")
            c.font = Font(b=True, size=14)
            # Set row heights to accommodate logo
            for logo_r in range(r, r + 2):
                ws.row_dimensions[logo_r].height = 30
            r += 1
            # Optional address line (simple placeholder to mimic layout)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
            ws.cell(row=r, column=2, value="Website: www.megascreen.com.my   Tel: 601-657 3233   Fax: 604-656 1318")
            r += 2

            # Only add metadata on the very first page
            if page_index == 1:
                # Quotation meta (box on the right)
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
                ws.cell(row=r, column=1, value=f"QUOTATION NO: {meta.get('quotation_id','')}").font = Font(b=True)
                r += 1
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
                ws.cell(row=r, column=1, value=f"Date: {_dt.date.today().strftime('%d-%b-%y')}")
                r += 1
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
                ws.cell(row=r, column=1, value=f"Page: {page_index} of {len(pages)}")
                r += 2
                # Title centered
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
                ws.cell(row=r, column=1, value="QUOTATION").font = Font(b=True, size=12)
                ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
                r += 2

                # Project/Remark boxes
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
                ws.cell(row=r, column=1, value=f"Project Name: {meta.get('project_name','')}")
                apply_range_border(r, 1, r, 5, border_all)
                ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=9)
                ws.cell(row=r, column=6, value=f"Modules Code: {meta.get('modules_code','')}")
                apply_range_border(r, 6, r, 9, border_all)
                r += 1
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
                ws.cell(row=r, column=1, value=f"Project Code: {meta.get('project_code','')}")
                apply_range_border(r, 1, r, 5, border_all)
                ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=9)
                ws.cell(row=r, column=6, value=f"Total Repair Modules : {meta.get('total_repair_modules','')}pcs")
                apply_range_border(r, 6, r, 9, border_all)
                r += 1

                # Date Request and Pixel row
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
                ws.cell(row=r, column=1, value="Date Request")
                apply_range_border(r, 1, r, 2, border_all)
                ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
                ws.cell(row=r, column=3, value=meta.get('date_request',''))
                apply_range_border(r, 3, r, 5, border_all)
                ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
                ws.cell(row=r, column=6, value="Pixel")
                apply_range_border(r, 6, r, 7, border_all)
                ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
                ws.cell(row=r, column=8, value=meta.get('pixel',''))
                apply_range_border(r, 8, r, 9, border_all)
                r += 2
            else:
                # For subsequent tables, skip the metadata block and go straight to table headers
                # Add quotation metadata (QUOTATION NO, Date, Page) to subsequent tables
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
                ws.cell(row=r, column=1, value=f"QUOTATION NO: {meta.get('quotation_id','')}")
                r += 1
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
                ws.cell(row=r, column=1, value=f"Date: {_dt.date.today().strftime('%d-%b-%y')}")
                r += 1
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
                ws.cell(row=r, column=1, value=f"Page: {page_index} of {len(pages)}")
                r += 2
                
                # Add "QUOTATION" title for consistency
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
                ws.cell(row=r, column=1, value="QUOTATION").font = Font(b=True, size=12)
                ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
                r += 2
                
                # Project metadata block (same as first table) for subsequent tables
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
                ws.cell(row=r, column=1, value=f"Project Name: {meta.get('project_name','')}")
                apply_range_border(r, 1, r, 5, border_all)
                ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=9)
                ws.cell(row=r, column=6, value=f"Modules Code: {meta.get('modules_code','')}")
                apply_range_border(r, 6, r, 9, border_all)
                r += 1
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
                ws.cell(row=r, column=1, value=f"Project Code: {meta.get('project_code','')}")
                apply_range_border(r, 1, r, 5, border_all)
                ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=9)
                ws.cell(row=r, column=6, value=f"Total Repair Modules : {meta.get('total_repair_modules','')}pcs")
                apply_range_border(r, 6, r, 9, border_all)
                r += 1

                # Date Request and Pixel row
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
                ws.cell(row=r, column=1, value="Date Request")
                apply_range_border(r, 1, r, 2, border_all)
                ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
                ws.cell(row=r, column=3, value=meta.get('date_request',''))
                apply_range_border(r, 3, r, 5, border_all)
                ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
                ws.cell(row=r, column=6, value="Pixel")
                apply_range_border(r, 6, r, 7, border_all)
                ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
                ws.cell(row=r, column=8, value=meta.get('pixel',''))
                apply_range_border(r, 8, r, 9, border_all)
                r += 2

            # Table headers styled with full borders
            issue_cols = list(issue_fields)
            # Build a synonym -> canonical issue header map (normalized)
            import re
            def _norm(s: str) -> str:
                return re.sub(r"[^a-z0-9]", "", (s or "").lower())
            ISSUE_SYNONYM_MAP = {}
            for hdr in issue_cols:
                ISSUE_SYNONYM_MAP[_norm(hdr)] = hdr
                for alt in ISSUE_KEY_MAP.get(hdr, []):
                    ISSUE_SYNONYM_MAP[_norm(alt)] = hdr
            def match_issue_name(text: str | None) -> str | None:
                if not text:
                    return None
                n = _norm(text)
                if n in ISSUE_SYNONYM_MAP:
                    return ISSUE_SYNONYM_MAP[n]
                # token/substring fallback
                for key, hdr in ISSUE_SYNONYM_MAP.items():
                    if key in n or n in key:
                        return hdr
                return None
            headers = ["Item", "Module No", "RN No"] + issue_cols + ["Quantity"]
            fill_grey = PatternFill("solid", fgColor="DDDDDD")
            start_col = 1
            for idx, h in enumerate(headers):
                col_idx = start_col + idx
                # Limit to column Q (17)
                if col_idx > 17:
                    break
                ws.cell(row=r, column=col_idx, value=h).font = Font(b=True, size=8)
                # Wrap long issue headers
                ws.cell(row=r, column=col_idx).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.cell(row=r, column=col_idx).fill = fill_grey
                ws.cell(row=r, column=col_idx).border = border_all
                if h in ("Module No"):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 7
                elif h in ("RN No"):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 8
                elif h == "Item":
                    ws.column_dimensions[get_column_letter(col_idx)].width = 5
                elif h == "Quantity":
                    ws.column_dimensions[get_column_letter(col_idx)].width = 8
                else:
                    # Issue columns
                    ws.column_dimensions[get_column_letter(col_idx)].width = 8
            # Slightly taller header row for readability
            ws.row_dimensions[r].height = 20
            r += 1

            # Stretch logo to approximate table width once column widths are known
            if logo_img is not None:
                try:
                    end_col = 3 + len(issue_cols) + 1
                    total_chars = 0.0
                    for c in range(1, end_col + 1):
                        w = ws.column_dimensions[get_column_letter(c)].width or 8
                        total_chars += float(w)
                    # Approximate pixel width from Excel character width
                    logo_img.width = int(total_chars * 7)
                except Exception:
                    pass

            # Table rows: always render 10 lines with full borders
            max_rows = page_size
            for i in range(max_rows):
                if i < len(page_rows):
                    v = page_rows[i]
                    item_no = (page_index-1)*page_size + i + 1
                    issue_val = str(v[3]).strip()
                    qty_val = v[4]
                else:
                    issue_val = ""
                    qty_val = ""
                # First three base columns
                # Module No column: use Project Code from meta instead of board module number
                module_cell = str(meta.get('project_code', '')).strip() if i < len(page_rows) else ""
                if not module_cell and i < len(page_rows):
                    module_cell = (page_rows[i][1] if i < len(page_rows) else "")
                base_vals = [item_no if i < len(page_rows) else "", module_cell, (page_rows[i][2] if i < len(page_rows) else "")]
                for off, val in enumerate(base_vals):
                    c = ws.cell(row=r, column=1+off, value=val)
                    c.border = border_all
                    c.alignment = align_center
                # Issue columns: prefer board counts if present (supports nested 'issues' dict); else use row Issue selection
                canon = match_issue_name(issue_val)
                # Build normalized board field map once per row
                b_norm_map = {}
                if i < len(page_rows):
                    bid_row = str(page_rows[i][0])
                    b_row = _get_board_by_id(bid_row) or {}
                    def _flatten(prefix, obj):
                        if isinstance(obj, dict):
                            for k, v in obj.items():
                                key = _norm((prefix + '_' + str(k)) if prefix else str(k))
                                b_norm_map[key] = v
                                _flatten(key, v)
                        # ignore lists; no expected structure
                    _flatten('', b_row)
                for j, issue_name in enumerate(issue_cols):
                    col_idx = 4 + j
                    # Try board-provided count first
                    val = None
                    aliases = ISSUE_KEY_MAP.get(issue_name, [])
                    # Also include the header itself as an alias
                    aliases = list(aliases) + [issue_name]
                    for alias in aliases:
                        nk = _norm(alias)
                        if nk in b_norm_map and b_norm_map.get(nk) not in (None, ""):
                            val = b_norm_map[nk]
                            break
                    # If still None, try any key containing the alias token
                    if val is None:
                        for bk, bv in b_norm_map.items():
                            if nk in bk and bv not in (None, ""):
                                val = bv
                                break
                    # Coerce numeric if possible
                    try:
                        val_num = int(val)
                    except Exception:
                        val_num = None
                    if val_num is not None and val_num != 0:
                        cell_val = val_num
                    else:
                        # Fall back to row Issue selection
                        try:
                            qn = int(qty_val)
                        except Exception:
                            qn = qty_val
                        cell_val = (qn if (canon == issue_name and qn not in (None, "", 0)) else "")
                    c = ws.cell(row=r, column=col_idx, value=cell_val)
                    c.border = border_all
                    c.font = Font(size=6.5)
                    c.alignment = align_center
                try:
                    # Final Quantity column
                    qv = int(qty_val) if str(qty_val).isdigit() else qty_val
                    c = ws.cell(row=r, column=3 + len(issue_cols) + 1, value=qv)
                    c.border = border_all
                    c.font = Font(size=6.5)
                    c.alignment = align_center
                except Exception:
                    c = ws.cell(row=r, column=3 + len(issue_cols) + 1, value=qty_val)
                    c.border = border_all
                    c.font = Font(size=6.5)
                    c.alignment = align_center
                # Reduce row height to fit more vertically
                ws.row_dimensions[r].height = 12
                r += 1
            # Totals row (Total Repair Modules) with full-width border
            try:
                total_qty = int(meta.get('total_repair_modules'))
            except Exception:
                total_qty = 0
                for v in rows:
                    try:
                        total_qty += int(v[4])
                    except Exception:
                        pass
            # Merge up to last column before Quantity
            end_merge_col = 3 + len(issue_cols)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=end_merge_col)
            ws.cell(row=r, column=1, value="Total Repair Modules (pcs)").font = Font(b=True)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="right")
            # Apply borders across merged range and qty cell
            for c in range(1, end_merge_col + 2):
                ws.cell(row=r, column=c).border = border_all
            ws.cell(row=r, column=end_merge_col + 1, value=total_qty).font = Font(b=True)
            r += 3

            # Textual Remark section (replicating provided sample)
            try:
                medium = Side(style="medium", color="000000")
                # "Remark:" label
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=17)
                ws.cell(row=r, column=1, value="Remark:").font = Font(b=True, size=10)
                ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
                r += 1
                # Notice text
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=17)
                ws.cell(row=r, column=1, value="** Please Notice that above information is just an estimate cost of repair & rework for Led Modules.")
                ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
                ws.row_dimensions[r].height = 18
                r += 2
                # Authorized by (left) and Date (right) on same row
                start_c = max(1, 17 - 6)
                end_c = max(start_c + 2, 17 - 2)
                left_end_col = max(1, start_c - 1)
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=left_end_col)
                ws.cell(row=r, column=1, value="Authorized  by :")
                ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
                # Date label aligned on right of the same row
                ws.merge_cells(start_row=r, start_column=start_c, end_row=r, end_column=end_c)
                ws.cell(row=r, column=start_c, value="Date:")
                ws.cell(row=r, column=start_c).alignment = Alignment(horizontal="right")
                r += 2
                # Signature line (top medium border across a few columns)
                sig_start_col = 1
                sig_end_col = max(4, min(6, 17))
                for cc in range(sig_start_col, sig_end_col + 1):
                    ws.cell(row=r, column=cc).border = Border(top=medium)
                ws.row_dimensions[r].height = 12
                r += 1
                # Team label
                ws.cell(row=r, column=1, value="Repair & Rework Team")
                ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
                r += 2
                # Spacer row after team line
                r += 1
            except Exception:
                pass
        wb.save(path)

    def export_to_csv(path: str, rows, meta: dict):
        import csv
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Quotation ID", meta.get('quotation_id','')])
            writer.writerow(["Project Name", meta.get('project_name','')])
            writer.writerow(["Project Code", meta.get('project_code','')])
            writer.writerow(["Modules Code", meta.get('modules_code','')])
            writer.writerow(["Total Repair Modules", meta.get('total_repair_modules','')])
            writer.writerow(["Date Request", meta.get('date_request','')])
            writer.writerow([])
            writer.writerow(["Item", "Module No", "RN No", "Issue", "Quantity"])
            for i, r in enumerate(rows, start=1):
                writer.writerow([i, r[1], r[2], r[3], r[4]])

    # Bindings
    btn_add.configure(command=add_selected_to_quote)
    btn_remove.configure(command=remove_selected_from_quote)
    btn_clear.configure(command=clear_quote)
    btn_export.configure(command=export_quote)
    ent_search.bind("<KeyRelease>", lambda _e: refresh_boards())

    # Checkbox toggle only on first column
    def on_tree_click(event):
        col = tv_boards.identify_column(event.x)
        row = tv_boards.identify_row(event.y)
        if col == "#1" and row:
            iid = row
            bid = iid.split(":", 1)[1] if ":" in iid else iid
            if bid in selected_ids:
                selected_ids.remove(bid)
            else:
                selected_ids.add(bid)
            vals = tv_boards.item(iid, "values")
            chk = "☑" if bid in selected_ids else "☐"
            tv_boards.item(iid, values=(chk,) + tuple(vals[1:]))
            return
    tv_boards.bind("<Button-1>", on_tree_click, add="+")

    refresh_boards()

    # Simple in-place editing for Issue and Quantity
    issue_fields = (
        "caterpillar",
        "pixel drop",
        "pixel problem",
        "kaki patah",
        "green/red/blue line",
        "box problem",
        "module blackout",
        "broken module",
        "broken connector",
        "broken power socket",
        "wiring",
        "broken frame",
    )
    # Mapping of issue header to possible board keys (normalized)
    ISSUE_KEY_MAP = {
        "caterpillar": ["caterpillar"],
        "pixel drop": ["lamp_pixel_drop", "pixel_drop"],
        "pixel problem": ["lamp_pixel_problem", "pixel_problem"],
        "kaki patah": ["kakipatah", "kaki_patah"],
        "green/red/blue line": [
            "green/red/blue line",
            "greenredblueline",
            "anomalyline",
            "line_issue",
            "rgb_line",
            "grb_line",
            "rgbline",
            "grbline",
        ],
        "box problem": ["boxproblem"],
        "module blackout": ["moduleblackout", "halfwholemoduleblackout"],
        "broken module": ["brokenmodule"],
        "broken connector": ["brokenconnector"],
        "broken power socket": ["brokenpowersocket"],
        "wiring": ["wiring"],
        "broken frame": ["brokenframe"],
    }

    def begin_edit(event):
        iid = tv_quote.identify_row(event.y)
        col = tv_quote.identify_column(event.x)
        if not iid or not col:
            return
        col_idx = int(col.replace('#','')) - 1
        bbox = tv_quote.bbox(iid, col)
        if not bbox:
            return
        x, y, w, h = bbox
        vals = list(tv_quote.item(iid, 'values'))
        # Issue column
        if q_cols[col_idx] == 'Issue':
            top = ttk.Combobox(tv_quote, values=issue_fields, state='readonly')
            top.place(x=x, y=y, width=w, height=h)
            top.set(vals[col_idx])
            def commit(_e=None):
                vals[col_idx] = top.get()
                tv_quote.item(iid, values=tuple(vals))
                top.destroy()
            top.bind('<Return>', commit)
            top.bind('<FocusOut>', commit)
            top.focus_set()
        # Quantity column
        elif q_cols[col_idx] == 'Quantity':
            var = tk.StringVar(value=str(vals[col_idx]))
            ent = ttk.Entry(tv_quote, textvariable=var)
            ent.place(x=x, y=y, width=w, height=h)
            def commit_q(_e=None):
                v = var.get().strip()
                try:
                    v = int(v)
                except Exception:
                    v = v
                vals[col_idx] = v
                tv_quote.item(iid, values=tuple(vals))
                ent.destroy()
            ent.bind('<Return>', commit_q)
            ent.bind('<FocusOut>', commit_q)
            ent.focus_set()

    tv_quote.bind('<Double-1>', begin_edit)

    return page
